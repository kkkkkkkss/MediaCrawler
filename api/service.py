# -*- coding: utf-8 -*-
# FastAPI 业务逻辑层
# 将 UrlCheckCrawler 的功能封装为可被 API 路由调用的异步函数

import os
import asyncio
from typing import List, Optional

import config
from api.schemas import MysqlSourceRequest, UrlCheckResult
from api.task_manager import TaskInfo
from store.url_check_excel_store import generate_url_check_excel, merge_results_to_excel
from tools import utils
from tools.url_detector import detect_platform, detect_source_platform, group_urls_by_platform
from tools.url_check_status import (
    STATUS_EXCEPTION,
    STATUS_INVALID,
    STATUS_UNSUPPORTED,
    STATUS_VALID,
    URL_CHECK_PLATFORM_ORDER,
    empty_metrics,
    is_supported_url_check_platform,
    should_clear_metrics,
    validity_label,
)
from tools.ai_field_mapper import ai_mapper
from tools.validity_checker import http_pre_check, check_api_json_validity, is_invalid


_PLATFORM_NAMES = {
    "dy": "抖音", "ks": "快手", "bili": "B站",
    "toutiao": "头条", "xigua": "西瓜视频", "wb": "微博", "unknown": "未知",
}


def _apply_single_status(result: UrlCheckResult, status_code: int, reason: str = ""):
    """单条响应也使用四态；特殊状态清空互动量，避免沿用旧字段。"""
    result.is_valid = status_code
    result.validity_label = validity_label(status_code)
    result.status_reason = reason
    if should_clear_metrics(status_code):
        result.praise_count = None
        result.reply_count = None
        result.visit_count = None
        result.share_count = None


def _result_status_text(result: dict) -> str:
    return validity_label(result.get("_is_valid"))


def _metric_display(metrics: dict, key: str) -> str:
    value = metrics.get(key)
    return "-" if value is None else str(value)


def _summarize_status_counts(results: list) -> str:
    counts = {STATUS_VALID: 0, STATUS_INVALID: 0, STATUS_UNSUPPORTED: 0, STATUS_EXCEPTION: 0}
    for r in results:
        code = r.get("_is_valid")
        if code in counts:
            counts[code] += 1
    return (
        f"有效{counts[STATUS_VALID]}, 无效{counts[STATUS_INVALID]}, "
        f"不支持{counts[STATUS_UNSUPPORTED]}, 异常{counts[STATUS_EXCEPTION]}"
    )


async def run_single_url_check(
    info: "TaskInfo",
    url: str,
    mode: str = "both",
    enable_comments: bool = False,
):
    """
    异步处理单条 URL（任务模式，日志实时推送到 TaskInfo）。
    完成后将 UrlCheckResult 存入 info.single_result。
    """
    platform, content_id = detect_platform(url)
    source_platform = detect_source_platform(url, platform)
    result = UrlCheckResult(id=1, url=url, platform=platform, source_platform=source_platform)
    info.add_log(
        f"平台识别: {_PLATFORM_NAMES.get(source_platform, _PLATFORM_NAMES.get(platform, platform))}, "
        f"检测链路={_PLATFORM_NAMES.get(platform, platform)}, content_id={content_id}"
    )

    if not is_supported_url_check_platform(platform):
        reason = f"平台 {platform or 'unknown'} 暂不支持 url_check，已跳过检测"
        _apply_single_status(result, STATUS_UNSUPPORTED, reason)
        result.content_type = _get_platform_type(platform)
        info.add_log(f"{reason} → 标记不支持")
        info.single_result = result.model_dump()
        return

    result.content_type = _get_platform_type(platform)

    # HTTP 预检
    info.add_log("执行 HTTP 预检...")
    status, final_url = await http_pre_check(url, platform)
    if is_invalid(status):
        _apply_single_status(result, STATUS_INVALID, f"HTTP 预检失效: {status.value}")
        info.add_log(f"HTTP 预检失效: {status.value} → 标记无效")
        info.single_result = result.model_dump()
        return
    info.add_log("HTTP 预检通过")

    # 浏览器阶段
    from playwright.async_api import async_playwright

    config.URLCHECK_MODE = mode
    config.URLCHECK_ENABLE_COMMENTS = enable_comments

    crawler = None
    async with async_playwright() as pw:
        try:
            from media_platform.url_check.core import UrlCheckCrawler
            crawler = UrlCheckCrawler()
            row = {
                "id": 1,
                "url": url,
                "_platform": platform,
                "_source_platform": source_platform,
                "_content_id": content_id,
            }
            info.add_log("正在启动浏览器客户端...")
            client, _ = await crawler._create_platform_client(platform, pw)
            if client:
                info.add_log("浏览器客户端创建成功，获取接口数据...")
                raw_json = await crawler._fetch_detail(platform, client, content_id, url, row=row)
                content_id = row.get("_content_id", content_id)

                if raw_json is not None:
                    api_status = check_api_json_validity(platform, raw_json)
                    if is_invalid(api_status):
                        _apply_single_status(result, STATUS_INVALID, f"接口字段检测失效: {api_status.value}")
                        info.add_log(f"接口字段检测失效: {api_status.value}")
                    else:
                        _apply_single_status(result, STATUS_VALID)
                        if mode != "validity":
                            if platform == "toutiao":
                                info.add_log("指标提取方式: DOM 直接提取")
                                # 西瓜已拿到内容 ID 时复用头条规范页，避免指标阶段又回到西瓜原页触发壳页/超时。
                                toutiao_orig_url = url if any(
                                    d in url for d in ("zjurl.cn", "weitoutiao")
                                ) else None
                                metrics = await client.get_article_metrics_from_dom(
                                    content_id, original_url=toutiao_orig_url
                                )
                            else:
                                async def _health_check(plat):
                                    from tools.platform_health_checker import check_benchmark
                                    return await check_benchmark(plat, client, crawler._fetch_detail)

                                metrics = await ai_mapper.extract_metrics(
                                    platform, raw_json,
                                    health_checker=_health_check,
                                    task_info=info,
                                )
                                info.add_log(f"指标提取方式: {ai_mapper.last_method or '硬编码'}")
                            result.praise_count = metrics.get("praise_count")
                            result.reply_count = metrics.get("reply_count")
                            result.visit_count = metrics.get("visit_count")
                            result.share_count = metrics.get("share_count")
                            result.author = metrics.get("author") or ""
                            info.add_log(
                                f"提取结果: 点赞={result.praise_count} 评论={result.reply_count} "
                                f"转发={result.share_count} 播放={result.visit_count}"
                            )

                        # 抓取评论
                        if enable_comments and content_id:
                            info.add_log(f"开始抓取评论 (最大 {config.URLCHECK_MAX_COMMENTS} 条)...")
                            comments_count = await crawler._fetch_and_store_comments(
                                platform, client, content_id, url, row
                            )
                            info.add_log(f"评论抓取完成，获取到 {comments_count} 条评论")
                            if comments_count and comments_count > 0:
                                if result.reply_count is None or result.reply_count == 0:
                                    result.reply_count = comments_count
                                    info.add_log(f"评论数已从实际抓取结果更新: {comments_count}")
                            # 收集评论到 TaskInfo
                            if row.get("_comments"):
                                info.comments_data = [{
                                    "content_url": url,
                                    "content_id": content_id,
                                    "platform": platform,
                                    "comments": row["_comments"],
                                }]
                else:
                    fail_reason = row.get("_fetch_fail_reason", "unknown")
                    if fail_reason == UrlCheckCrawler._FETCH_FAIL_CONTENT:
                        _apply_single_status(result, STATUS_INVALID, "内容不存在或已删除")
                        info.add_log(f"接口获取失败: {fail_reason} → 标记无效")
                    else:
                        _apply_single_status(
                            result,
                            STATUS_EXCEPTION,
                            f"接口获取失败，疑似风控/网络/浏览器异常: {fail_reason}",
                        )
                        info.add_log(f"接口获取失败: {fail_reason} → 标记检测异常")
            else:
                _apply_single_status(result, STATUS_EXCEPTION, "浏览器客户端创建失败")
                info.add_log("浏览器客户端创建失败 → 标记检测异常")
        except Exception as e:
            utils.logger.error(f"[API] 单链接处理异常: {e}")
            _apply_single_status(result, STATUS_EXCEPTION, f"处理异常: {e}")
            info.add_log(f"处理异常: {e} → 标记检测异常")
        finally:
            if crawler:
                # 旧逻辑只在正常路径清理；现在异常/取消路径也收口，避免后台残留无头浏览器。
                await crawler._cleanup_browser()

    info.single_result = result.model_dump()
    info.add_log("单条检测完成")


async def run_batch_check(
    info: TaskInfo,
    urls: List[str],
    mode: str = "both",
    enable_comments: bool = False,
):
    """批量URL检测任务（后台异步执行，多平台并行处理）"""
    config.URLCHECK_MODE = mode
    config.URLCHECK_ENABLE_COMMENTS = enable_comments
    config.URLCHECK_INPUT_SOURCE = "file"

    rows = []
    for idx, url in enumerate(urls, start=1):
        url = url.strip()
        if url.startswith("http"):
            rows.append({"id": idx, "url": url})

    info.total = len(rows)
    if not rows:
        info.status = "completed"
        info.message = "没有有效的URL"
        return

    from media_platform.url_check.core import UrlCheckCrawler
    groups = group_urls_by_platform(rows)

    platform_order = URL_CHECK_PLATFORM_ORDER
    active_platforms = [p for p in platform_order if p in groups]

    # 多平台并行处理：每个平台独立浏览器，互不干扰
    parallel_enabled = getattr(config, "URLCHECK_PARALLEL_PLATFORMS", True)
    if parallel_enabled and len(active_platforms) > 1:
        info.add_log(
            f"启用多平台并行模式，同时处理 {len(active_platforms)} 个平台: "
            f"{[_PLATFORM_NAMES.get(p, p) for p in active_platforms]}"
        )
        all_results = await _parallel_platform_process(
            info, groups, active_platforms, mode, _PLATFORM_NAMES
        )
    else:
        all_results = await _sequential_platform_process(
            info, groups, active_platforms, mode, _PLATFORM_NAMES
        )

    await _append_unsupported_results(info, groups, all_results)

    # 保存结果数据到 TaskInfo
    if all_results:
        info.result_data = all_results
        info.comments_data = _collect_comments_data(all_results)

        import pathlib
        output_dir = pathlib.Path("data/url_check/excel")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(output_dir / f"{info.task_id}.xlsx")
        excel_path = generate_url_check_excel(all_results, output_path)
        info.result_file = excel_path
        info.add_log(f"Excel 报表已生成: {os.path.basename(excel_path)}")

    info.progress = 100.0


async def run_file_upload_check(
    info: TaskInfo,
    file_path: str,
    url_column: str,
    mode: str,
    enable_comments: bool,
):
    """从上传的 Excel/CSV 文件中提取 URL 并检测，多平台并行处理"""
    import pathlib

    ext = os.path.splitext(file_path)[1].lower()
    is_excel = ext in (".xlsx", ".xls")

    urls = _extract_urls_from_file(file_path, url_column)
    if not urls:
        info.status = "failed"
        info.message = "文件中未找到有效URL"
        return

    config.URLCHECK_MODE = mode
    config.URLCHECK_ENABLE_COMMENTS = enable_comments
    config.URLCHECK_INPUT_SOURCE = "file"

    rows = []
    for idx, url in enumerate(urls, start=1):
        url = url.strip()
        if url.startswith("http"):
            rows.append({"id": idx, "url": url})

    info.total = len(rows)
    if not rows:
        info.status = "completed"
        info.message = "没有有效的URL"
        return

    from media_platform.url_check.core import UrlCheckCrawler
    groups = group_urls_by_platform(rows)

    platform_order = URL_CHECK_PLATFORM_ORDER
    active_platforms = [p for p in platform_order if p in groups]
    info.add_log(f"文件解析完成，共 {len(rows)} 条URL待检测")

    # 多平台并行处理
    parallel_enabled = getattr(config, "URLCHECK_PARALLEL_PLATFORMS", True)
    if parallel_enabled and len(active_platforms) > 1:
        info.add_log(
            f"启用多平台并行模式，同时处理 {len(active_platforms)} 个平台: "
            f"{[_PLATFORM_NAMES.get(p, p) for p in active_platforms]}"
        )
        all_results = await _parallel_platform_process(
            info, groups, active_platforms, mode, _PLATFORM_NAMES
        )
    else:
        all_results = await _sequential_platform_process(
            info, groups, active_platforms, mode, _PLATFORM_NAMES
        )

    await _append_unsupported_results(info, groups, all_results)

    if all_results:
        info.result_data = all_results
        info.comments_data = _collect_comments_data(all_results)

    output_dir = pathlib.Path("data/url_check/excel")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = str(output_dir / f"{info.task_id}.xlsx")

    if is_excel:
        excel_path = merge_results_to_excel(
            source_path=file_path,
            results=all_results,
            output_path=output_path,
            url_column=url_column,
        )
        info.result_file = excel_path
        info.add_log(f"Excel 报表已生成: {os.path.basename(excel_path)}")
    elif all_results:
        excel_path = generate_url_check_excel(all_results, output_path)
        info.result_file = excel_path
        info.add_log(f"Excel 报表已生成: {os.path.basename(excel_path)}")

    info.progress = 100.0

    try:
        os.unlink(file_path)
    except Exception:
        pass


async def run_mysql_check(info: TaskInfo, req: MysqlSourceRequest):
    """从外部 MySQL 读取 URL 并检测，结果回写同一数据库"""
    import aiomysql

    try:
        pool = await aiomysql.create_pool(
            host=req.host, port=req.port,
            user=req.user, password=req.password,
            db=req.database, charset="utf8mb4",
            autocommit=True, minsize=1, maxsize=3,
        )
        async with pool.acquire() as conn:
            async with conn.cursor(aiomysql.DictCursor) as cur:
                sql = f"SELECT * FROM {req.table} WHERE {req.url_column} IS NOT NULL AND {req.url_column} != '' LIMIT %s"
                await cur.execute(sql, (req.batch_size,))
                rows = await cur.fetchall()

        pool.close()
        await pool.wait_closed()

        if not rows:
            info.status = "completed"
            info.message = "数据库中无待处理URL"
            return

        await _run_db_rows_check(info, rows, req.url_column, req.mode.value, req.enable_comments)
    except Exception as e:
        info.status = "failed"
        info.message = f"数据库连接失败: {e}"


async def _run_db_rows_check(
    info: TaskInfo,
    db_rows: List[dict],
    url_column: str,
    mode: str = "both",
    enable_comments: bool = False,
):
    """处理从 DB 读取的行（保留原始 ID，结果回写 DB），多平台并行处理"""
    config.URLCHECK_MODE = mode
    config.URLCHECK_ENABLE_COMMENTS = enable_comments
    config.URLCHECK_INPUT_SOURCE = "db"

    from database.external_db import external_db
    await external_db.ensure_pool()

    url_rows = []
    for r in db_rows:
        url = r.get(url_column, "")
        if url and str(url).startswith("http"):
            url_rows.append({"id": r["id"], "url": str(url).strip()})

    info.total = len(url_rows)
    if not url_rows:
        info.status = "completed"
        info.message = "没有有效的URL"
        return

    from media_platform.url_check.core import UrlCheckCrawler
    groups = group_urls_by_platform(url_rows)

    platform_order = URL_CHECK_PLATFORM_ORDER
    active_platforms = [p for p in platform_order if p in groups]
    info.add_log(f"数据库读取完成，共 {len(url_rows)} 条URL待检测")

    # 多平台并行处理
    parallel_enabled = getattr(config, "URLCHECK_PARALLEL_PLATFORMS", True)
    if parallel_enabled and len(active_platforms) > 1:
        info.add_log(
            f"启用多平台并行模式，同时处理 {len(active_platforms)} 个平台: "
            f"{[_PLATFORM_NAMES.get(p, p) for p in active_platforms]}"
        )
        all_results = await _parallel_platform_process(
            info, groups, active_platforms, mode, _PLATFORM_NAMES
        )
    else:
        all_results = await _sequential_platform_process(
            info, groups, active_platforms, mode, _PLATFORM_NAMES
        )

    await _append_unsupported_results(info, groups, all_results)

    if all_results:
        info.result_data = all_results
        info.comments_data = _collect_comments_data(all_results)

        import pathlib
        output_dir = pathlib.Path("data/url_check/excel")
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(output_dir / f"{info.task_id}.xlsx")
        excel_path = generate_url_check_excel(all_results, output_path)
        info.result_file = excel_path

    info.progress = 100.0


async def _append_unsupported_results(info: TaskInfo, groups: dict, all_results: list):
    """将未接入 url_check 的平台直接写为不支持，不再进入 HTTP/浏览器检测。"""
    from media_platform.url_check.core import UrlCheckCrawler

    unsupported = [
        (platform, rows)
        for platform, rows in groups.items()
        if not is_supported_url_check_platform(platform)
    ]
    if not unsupported:
        return

    crawler = UrlCheckCrawler()
    for platform, rows in unsupported:
        for row in rows:
            display_platform = row.get("_source_platform") or platform
            pname = _PLATFORM_NAMES.get(display_platform, _PLATFORM_NAMES.get(platform, platform or "未知"))
            reason = f"平台 {pname} 暂不支持 url_check，已跳过检测"
            await crawler._save_result(
                row, is_valid=STATUS_UNSUPPORTED, metrics=empty_metrics(), reason=reason
            )
            info.add_log(f"  [{pname}] 不支持 | {row.get('url', '')[:50]}")
            info.processed += 1
            if info.total:
                info.progress = min((info.processed / info.total) * 100, 99.9)

    all_results.extend(crawler._all_results)


async def _parallel_platform_process(
    info: TaskInfo,
    groups: dict,
    active_platforms: List[str],
    mode: str,
    platform_names: dict,
) -> list:
    """
    多平台并行处理核心逻辑：
    为每个平台创建独立的 Crawler 实例，通过 asyncio.gather 并行执行。
    每个平台内部由 _process_platform 自动判断是否多浏览器并发。
    """
    from media_platform.url_check.core import UrlCheckCrawler

    async def _platform_worker(platform: str):
        """单平台工作协程：委托给 _process_platform（内部自动多浏览器并发）"""
        url_rows = groups[platform]
        crawler = UrlCheckCrawler()
        pname = platform_names.get(platform, platform)

        concurrency = crawler._resolve_concurrency(platform, len(url_rows))
        info.add_log(
            f"[{pname}] 待处理 {len(url_rows)} 条, 浏览器并发={concurrency}"
        )

        # 实时回调：每处理完一条URL就立即推送日志并更新进度
        def _on_result(r: dict):
            url_short = r.get("url", "")[:50]
            valid_str = _result_status_text(r)
            metrics = r.get("_metrics", {})
            method = r.get("_extract_method", "-")
            display_platform = r.get("_source_platform") or platform
            display_pname = platform_names.get(display_platform, pname)
            # 展示 Worker 编号和 Cookie 标识
            w_id = r.get("_worker_id")
            c_id = r.get("_cookie_id")
            worker_tag = f"W{w_id}" if w_id else "W1"
            cookie_tag = f"C{c_id}" if c_id else "无Cookie"
            info.add_log(
                f"  [{display_pname}][{worker_tag}|{cookie_tag}][{method}] {valid_str} | "
                f"赞={_metric_display(metrics, 'praise_count')} "
                f"评={_metric_display(metrics, 'reply_count')} "
                f"转={_metric_display(metrics, 'share_count')} "
                f"播={_metric_display(metrics, 'visit_count')} | "
                f"{url_short}"
            )
            info.processed += 1
            info.progress = min((info.processed / info.total) * 100, 99.9)

        try:
            await crawler._process_platform(platform, url_rows, mode, on_result=_on_result)
        except Exception as e:
            info.add_log(f"[{pname}] 处理异常: {e}")
            utils.logger.error(f"[API-Parallel] 平台 {platform} 异常: {e}")

        result_count = len(crawler._all_results)
        info.add_log(
            f"[{pname}] 完成: 共{result_count}条, {_summarize_status_counts(crawler._all_results)}"
        )
        return crawler

    # 并行启动所有平台
    tasks = [_platform_worker(p) for p in active_platforms]
    workers = await asyncio.gather(*tasks, return_exceptions=True)

    # 合并所有平台的结果
    all_results = []
    for result in workers:
        if isinstance(result, Exception):
            utils.logger.error(f"[API-Parallel] worker 异常: {result}")
            continue
        if hasattr(result, "_all_results"):
            all_results.extend(result._all_results)

    return all_results


async def _sequential_platform_process(
    info: TaskInfo,
    groups: dict,
    active_platforms: List[str],
    mode: str,
    platform_names: dict,
) -> list:
    """顺序处理模式（单平台或未启用并行时的回退逻辑）"""
    from media_platform.url_check.core import UrlCheckCrawler

    crawler = UrlCheckCrawler()

    for platform in active_platforms:
        url_rows = groups[platform]
        pname = platform_names.get(platform, platform)
        concurrency = crawler._resolve_concurrency(platform, len(url_rows))
        info.add_log(
            f"开始处理平台: {pname}，共 {len(url_rows)} 条, 浏览器并发={concurrency}"
        )

        if info.is_cancelled:
            info.add_log("任务已被用户终止")
            return crawler._all_results

        # 实时回调
        def _on_result(r: dict, _pname=pname):
            url_short = r.get("url", "")[:50]
            valid_str = _result_status_text(r)
            metrics = r.get("_metrics", {})
            method = r.get("_extract_method", "-")
            display_platform = r.get("_source_platform") or platform
            display_pname = platform_names.get(display_platform, _pname)
            w_id = r.get("_worker_id")
            c_id = r.get("_cookie_id")
            worker_tag = f"W{w_id}" if w_id else "W1"
            cookie_tag = f"C{c_id}" if c_id else "无Cookie"
            info.add_log(
                f"  [{display_pname}][{worker_tag}|{cookie_tag}][{method}] {valid_str} | "
                f"赞={_metric_display(metrics, 'praise_count')} "
                f"评={_metric_display(metrics, 'reply_count')} "
                f"转={_metric_display(metrics, 'share_count')} "
                f"播={_metric_display(metrics, 'visit_count')} | "
                f"{url_short}"
            )
            info.processed += 1
            info.progress = min((info.processed / info.total) * 100, 99.9)

        prev_count = len(crawler._all_results)
        try:
            await crawler._process_platform(platform, url_rows, mode, on_result=_on_result)
        except Exception as e:
            info.add_log(f"平台 {platform} 处理异常: {e}")
            utils.logger.error(f"[API] 批量处理平台 {platform} 异常: {e}")

        new_results = crawler._all_results[prev_count:]
        info.add_log(
            f"[{pname}] 完成: 共{len(new_results)}条, {_summarize_status_counts(new_results)}"
        )

    return crawler._all_results


def _extract_urls_from_file(file_path: str, url_column: str) -> List[str]:
    """从 Excel/CSV/TXT 中提取 URL 列表"""
    ext = os.path.splitext(file_path)[1].lower()
    urls = []

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            col_idx = None
            for i, h in enumerate(headers):
                if h and str(h).strip().lower() == url_column.lower():
                    col_idx = i
                    break
            # 模糊匹配：包含"链接"/"url"/"地址"的列
            if col_idx is None:
                for i, h in enumerate(headers):
                    hs = str(h) if h else ""
                    if "链接" in hs or "url" in hs.lower() or "地址" in hs:
                        col_idx = i
                        break
            if col_idx is not None:
                for row in ws.iter_rows(min_row=2):
                    val = row[col_idx].value
                    if val and str(val).startswith("http"):
                        urls.append(str(val).strip())
            wb.close()
        except Exception as e:
            utils.logger.error(f"[API] Excel 读取失败: {e}")

    elif ext == ".csv":
        import csv
        try:
            with open(file_path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    val = row.get(url_column, "")
                    if val and val.startswith("http"):
                        urls.append(val.strip())
        except Exception as e:
            utils.logger.error(f"[API] CSV 读取失败: {e}")

    elif ext == ".txt":
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("http"):
                        urls.append(line)
        except Exception as e:
            utils.logger.error(f"[API] TXT 读取失败: {e}")

    return urls


def _get_platform_type(platform: str) -> str:
    types = {"dy": "视频", "ks": "视频", "bili": "视频", "toutiao": "视频", "xhs": "笔记", "wb": "微博"}
    return types.get(platform, "未知")


def _collect_comments_data(results: list) -> list:
    """从 _all_results 中提取评论数据，按作品分组"""
    comments_data = []
    for r in results:
        comments = r.get("_comments")
        if comments:
            comments_data.append({
                "content_url": r.get("url", ""),
                "content_id": r.get("_content_id", ""),
                "platform": r.get("_platform", ""),
                "comments": comments,
            })
    return comments_data
