# -*- coding: utf-8 -*-
# 举报投诉 API 路由
# 所有 /api/v1/report 下的端点

import os
import tempfile

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse

from api.schemas_report import (
    ReportBatchRequest,
    ReportMysqlRequest,
    ReportProgressResponse,
    ReportReasonResponse,
    ReportResultItem,
    ReportSingleRequest,
    ReportTaskResponse,
    ReportTaskResultResponse,
)
from api.task_manager import TaskInfo, task_manager
from media_platform.report.report_config import (
    DEFAULT_REASON,
    PLATFORM_NAMES,
    PLATFORM_REPORT_REASONS,
    SUPPORTED_PLATFORMS,
)
from media_platform.report.report_engine import (
    generate_report_excel,
    get_latest_screenshot_base64,
    get_screenshot_base64,
    run_report_task,
)
from tools.url_detector import extract_urls_from_text

report_router = APIRouter(prefix="/api/v1/report", tags=["举报投诉"])


# ════════════════════ 举报理由查询 ════════════════════

@report_router.get("/reasons", response_model=ReportReasonResponse, summary="获取平台举报理由列表")
async def get_report_reasons(platform: str = "dy"):
    """
    获取指定平台可用的举报理由列表。
    前端据此动态渲染理由下拉选项。
    """
    if platform not in PLATFORM_REPORT_REASONS:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的平台: {platform}，可选: {SUPPORTED_PLATFORMS}"
        )
    reasons = list(PLATFORM_REPORT_REASONS[platform].keys())
    return ReportReasonResponse(
        platform=platform,
        platform_name=PLATFORM_NAMES.get(platform, platform),
        reasons=reasons,
        default_reason=DEFAULT_REASON.get(platform, reasons[0]),
    )


@report_router.get("/reasons/all", summary="获取所有平台举报理由")
async def get_all_report_reasons():
    """返回所有平台的举报理由映射（前端可一次加载缓存）"""
    result = {}
    for plat, reasons in PLATFORM_REPORT_REASONS.items():
        result[plat] = {
            "platform_name": PLATFORM_NAMES.get(plat, plat),
            "reasons": list(reasons.keys()),
            "default_reason": DEFAULT_REASON.get(plat, ""),
        }
    return JSONResponse(result)


# ════════════════════ 举报任务提交 ════════════════════

@report_router.post("/single", response_model=ReportTaskResponse, summary="单条链接举报")
async def report_single(req: ReportSingleRequest):
    """
    提交单条链接举报任务。

    支持直接URL或含文字的分享文本，后端自动提取链接。
    使用该平台所有可用Cookie各举报1次。
    """
    urls = extract_urls_from_text(req.url)
    if not urls:
        raise HTTPException(status_code=400, detail="未从输入中提取到有效链接")

    async def task_coro(info: TaskInfo):
        await run_report_task(info, urls, req.reason, req.description)

    task_id = task_manager.submit(task_coro, total=0, prefix="report-single")
    return ReportTaskResponse(
        task_id=task_id, status="pending",
        message=f"举报任务已提交，提取到 {len(urls)} 条链接",
        url_count=len(urls),
    )


@report_router.post("/batch", response_model=ReportTaskResponse, summary="批量链接举报")
async def report_batch(req: ReportBatchRequest):
    """
    提交批量链接举报任务。

    每条输入可以是完整URL或含文字的分享文本，后端自动提取链接。
    所有链接使用该平台所有可用Cookie各举报1次。
    """
    all_urls = []
    for text in req.urls:
        extracted = extract_urls_from_text(text)
        all_urls.extend(extracted)

    # 去重保序
    seen = set()
    unique_urls = []
    for u in all_urls:
        if u not in seen:
            seen.add(u)
            unique_urls.append(u)

    if not unique_urls:
        raise HTTPException(status_code=400, detail="未从输入中提取到有效链接")

    async def task_coro(info: TaskInfo):
        await run_report_task(info, unique_urls, req.reason, req.description)

    task_id = task_manager.submit(task_coro, total=0, prefix="report-batch")
    return ReportTaskResponse(
        task_id=task_id, status="pending",
        message=f"批量举报任务已提交，共 {len(unique_urls)} 条链接",
        url_count=len(unique_urls),
    )


@report_router.post("/upload", response_model=ReportTaskResponse, summary="上传文件举报")
async def report_upload(
    file: UploadFile = File(..., description="上传的文件（支持 .xlsx/.csv/.txt）"),
    url_column: str = Form(default="url", description="文件中存放URL的列名"),
    reason: str = Form(default="不实信息", description="举报理由"),
    description: str = Form(default="", description="补充说明（可选）"),
):
    """
    上传文件进行批量举报。

    支持格式：
    - .txt：每行一个URL或分享文本
    - .xlsx/.csv：读取指定列的URL
    """
    suffix = os.path.splitext(file.filename or "upload")[1] or ".txt"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix, dir="data/url_check")
    content = await file.read()
    tmp.write(content)
    tmp.close()

    # 从文件解析 URL 列表
    urls = _parse_urls_from_file(tmp.name, url_column)
    os.unlink(tmp.name)

    if not urls:
        raise HTTPException(status_code=400, detail="文件中未找到有效链接")

    async def task_coro(info: TaskInfo):
        await run_report_task(info, urls, reason, description)

    fname = os.path.splitext(file.filename or "file")[0][:10]
    task_id = task_manager.submit(task_coro, total=0, prefix=f"report-{fname}")
    return ReportTaskResponse(
        task_id=task_id, status="pending",
        message=f"文件举报任务已提交，解析到 {len(urls)} 条链接",
        url_count=len(urls),
    )


@report_router.post("/mysql", response_model=ReportTaskResponse, summary="从MySQL读取链接举报")
async def report_mysql(req: ReportMysqlRequest):
    """
    从指定 MySQL 表中读取URL进行批量举报。
    复用现有链接检测的数据库连接逻辑。
    """
    async def task_coro(info: TaskInfo):
        urls = await _read_urls_from_mysql(req)
        if not urls:
            info.add_log("数据库中未读取到有效链接")
            info.message = "无链接可举报"
            return
        info.add_log(f"从数据库读取 {len(urls)} 条链接")
        await run_report_task(info, urls, req.reason, req.description)

    task_id = task_manager.submit(task_coro, total=0, prefix="report-mysql")
    return ReportTaskResponse(
        task_id=task_id, status="pending",
        message="MySQL举报任务已提交",
    )


# ════════════════════ 任务进度/结果 ════════════════════

@report_router.get("/{task_id}", response_model=ReportProgressResponse, summary="查询举报任务进度")
async def get_report_progress(task_id: str, log_offset: int = 0):
    """查询举报任务的进度、日志和最新截图预览。"""
    info = task_manager.get_task(task_id)
    if not info:
        raise HTTPException(status_code=404, detail="任务不存在")

    logs = info.logs[log_offset:] if log_offset < len(info.logs) else []

    # 获取最新截图 base64（仅运行中时获取，减少IO）
    latest_ss = None
    if info.status == "running":
        latest_ss = get_latest_screenshot_base64(task_id)

    return ReportProgressResponse(
        task_id=info.task_id,
        status=info.status,
        progress=info.progress,
        total=info.total,
        processed=info.processed,
        message=info.message,
        logs=logs,
        log_total=len(info.logs),
        latest_screenshot=latest_ss,
    )


@report_router.get("/{task_id}/screenshots", summary="下载截图ZIP包")
async def download_report_screenshots(task_id: str):
    """下载举报任务的所有截图（ZIP打包）。"""
    zip_path = os.path.join("data", "report_screenshots", f"{task_id}.zip")
    if not os.path.exists(zip_path):
        raise HTTPException(status_code=404, detail="截图文件不存在（任务可能尚未完成）")
    return FileResponse(
        zip_path,
        media_type="application/zip",
        filename=f"report_screenshots_{task_id}.zip",
    )


@report_router.get("/{task_id}/screenshots/latest", summary="获取最新截图预览")
async def get_latest_screenshot(task_id: str):
    """获取最新一张截图的 base64 编码（前端实时预览用）。"""
    b64 = get_latest_screenshot_base64(task_id)
    if not b64:
        return JSONResponse({"screenshot": None, "message": "暂无截图"})
    return JSONResponse({"screenshot": b64})


@report_router.get("/{task_id}/screenshot/{filename}", summary="获取指定截图")
async def get_single_screenshot(task_id: str, filename: str):
    """获取指定截图文件的 base64 编码（结果表格展开行截图预览用）。"""
    b64 = get_screenshot_base64(task_id, filename)
    if not b64:
        raise HTTPException(status_code=404, detail="截图不存在")
    return JSONResponse({"screenshot": b64})


@report_router.get("/{task_id}/excel", summary="下载举报结果Excel")
async def download_report_excel(task_id: str):
    """生成并下载举报结果的 Excel 报告。"""
    info = task_manager.get_task(task_id)
    if not info:
        raise HTTPException(status_code=404, detail="任务不存在")
    if info.status != "completed":
        raise HTTPException(status_code=400, detail=f"任务状态: {info.status}，尚未完成")

    results = info.result_data or []
    if not results:
        raise HTTPException(status_code=400, detail="无举报结果可导出")

    excel_path = generate_report_excel(task_id, results)
    if not excel_path or not os.path.exists(excel_path):
        raise HTTPException(status_code=500, detail="Excel 生成失败")

    return FileResponse(
        excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"举报结果_{task_id}.xlsx",
    )


@report_router.get("/{task_id}/result", summary="获取举报结果")
async def get_report_result(task_id: str):
    """获取举报任务的详细结果（JSON格式）。"""
    info = task_manager.get_task(task_id)
    if not info:
        raise HTTPException(status_code=404, detail="任务不存在")
    if info.status != "completed":
        raise HTTPException(status_code=400, detail=f"任务状态: {info.status}，尚未完成")

    results = info.result_data or []
    success_count = sum(1 for r in results if r.get("success"))
    fail_count = len(results) - success_count

    return ReportTaskResultResponse(
        task_id=task_id,
        total=len(results),
        success_count=success_count,
        fail_count=fail_count,
        results=[ReportResultItem(**r) for r in results],
    )


@report_router.post("/{task_id}/cancel", summary="取消举报任务")
async def cancel_report_task(task_id: str):
    """取消正在执行的举报任务。"""
    info = task_manager.get_task(task_id)
    if not info:
        raise HTTPException(status_code=404, detail="任务不存在")
    if info.status in ("completed", "failed", "cancelled"):
        return JSONResponse({"success": False, "message": f"任务已处于终态: {info.status}"})
    info.cancel()
    return JSONResponse({"success": True, "message": "举报任务已标记为取消"})


# ════════════════════ 辅助函数 ════════════════════

def _parse_urls_from_file(filepath: str, url_column: str = "url") -> list[str]:
    """从文件中解析 URL 列表"""
    ext = os.path.splitext(filepath)[1].lower()
    urls = []

    if ext == ".txt":
        with open(filepath, "r", encoding="utf-8") as f:
            for line in f:
                extracted = extract_urls_from_text(line.strip())
                urls.extend(extracted)

    elif ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active
        # 找到 URL 列的索引
        header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = None
        for i, h in enumerate(header):
            if h == url_column.lower() or "url" in h or "链接" in h or "link" in h:
                col_idx = i
                break
        if col_idx is None:
            col_idx = 0  # 默认第一列

        for row in ws.iter_rows(min_row=2, values_only=True):
            if col_idx < len(row) and row[col_idx]:
                extracted = extract_urls_from_text(str(row[col_idx]))
                urls.extend(extracted)
        wb.close()

    elif ext == ".csv":
        import csv
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                val = row.get(url_column, "") or ""
                if not val:
                    # 尝试匹配包含 url 的列
                    for k, v in row.items():
                        if "url" in k.lower() or "链接" in k.lower():
                            val = v
                            break
                extracted = extract_urls_from_text(val)
                urls.extend(extracted)

    # 去重保序
    seen = set()
    unique = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            unique.append(u)
    return unique


async def _read_urls_from_mysql(req: ReportMysqlRequest) -> list[str]:
    """从 MySQL 读取 URL 列表"""
    import aiomysql
    urls = []
    try:
        pool = await aiomysql.create_pool(
            host=req.host, port=req.port,
            user=req.user, password=req.password,
            db=req.database, charset="utf8mb4",
            minsize=1, maxsize=2,
        )
        where_clause = f"WHERE {req.where}" if req.where else ""
        sql = f"SELECT `{req.url_column}` FROM `{req.table}` {where_clause} LIMIT {req.limit}"
        async with pool.acquire() as conn:
            async with conn.cursor() as cur:
                await cur.execute(sql)
                rows = await cur.fetchall()
                for row in rows:
                    if row[0]:
                        extracted = extract_urls_from_text(str(row[0]))
                        urls.extend(extracted)
        pool.close()
        await pool.wait_closed()
    except Exception as e:
        from tools import utils
        utils.logger.error(f"[ReportMySQL] 读取数据库失败: {e}")

    # 去重保序
    seen = set()
    unique = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            unique.append(u)
    return unique
