# -*- coding: utf-8 -*-
# 举报调度引擎
# 负责：URL 解析 -> 平台路由 -> Cookie 分配 -> 并发/串行控制 -> 截图管理 -> DB 写入 -> Excel 导出 -> 进度上报

import asyncio
import base64
import os
import random
import zipfile
from typing import List, Optional

import config
from media_platform.report.base_report import BaseReport, ReportResult
from media_platform.report.douyin_report import DouyinReport
from media_platform.report.kuaishou_report import KuaishouReport
from media_platform.report.report_config import DEFAULT_REASON, PLATFORM_NAMES, SUPPORTED_PLATFORMS
from media_platform.report.toutiao_report import ToutiaoReport
from media_platform.report.weibo_report import WeiboReport
from proxy.cookie_pool import cookie_pool
from tools import utils
from tools.url_detector import detect_platform

_REPORT_HANDLERS: dict[str, BaseReport] = {
    "dy": DouyinReport(),
    "wb": WeiboReport(),
    "ks": KuaishouReport(),
    "toutiao": ToutiaoReport(),
}


async def run_report_task(
    task_info,
    urls: List[str],
    reason: str,
    description: str = "",
):
    """
    举报任务主入口，由 API 层的 TaskManager 调用。
    """
    task_id = task_info.task_id
    all_results: List[ReportResult] = []

    platform_urls = _group_by_platform(urls)
    unsupported = platform_urls.pop("unknown", [])
    if unsupported:
        for u in unsupported:
            task_info.add_log(f"[跳过] 不支持的平台链接: {u}")

    total_ops = 0
    for platform, url_list in platform_urls.items():
        cookie_count = _get_cookie_count(platform)
        total_ops += len(url_list) * max(cookie_count, 1)
    task_info.total = total_ops
    task_info.add_log(f"任务开始: {len(urls)} 条链接, 预计 {total_ops} 次举报操作")

    max_concurrency = getattr(config, "REPORT_MAX_CONCURRENCY", 3)
    semaphore = asyncio.Semaphore(max_concurrency)

    # 多平台并行：每个平台独立的协程同时执行，共享信号量限制总并发浏览器数
    # 对同一账号来说，每次仍只处理一条链接（由信号量保证）
    processed_lock = asyncio.Lock()
    processed_counter = {"n": 0}

    async def _process_platform(platform: str, url_list: List[str]):
        """单个平台的处理协程，多平台实例并行运行"""
        handler = _REPORT_HANDLERS.get(platform)
        if not handler:
            for u in url_list:
                task_info.add_log(f"[跳过] 平台 {platform} 暂不支持举报: {u}")
            return []

        plat_name = PLATFORM_NAMES.get(platform, platform)
        task_info.add_log(f"--- 开始处理 {plat_name} 平台 ({len(url_list)} 条链接) ---")
        reason_text = _resolve_reason(platform, reason)
        plat_results: List[ReportResult] = []

        for url in url_list:
            if task_info.is_cancelled:
                task_info.add_log(f"[{plat_name}] 任务已取消")
                break

            cookies = _get_all_cookies(platform)
            if not cookies:
                cookies = [("guest", "")]

            task_info.add_log(f"[{plat_name}] 举报链接: {url} (将使用 {len(cookies)} 个账号)")

            # 先用第一个 Cookie 检测链接是否有效（避免失效链接浪费所有Cookie的时间）
            first_cid, first_cstr = cookies[0]
            first_result = await _single_report_with_semaphore(
                semaphore, handler, url, first_cstr, first_cid,
                reason_text, description, task_id, task_info, plat_name,
            )
            if isinstance(first_result, ReportResult):
                plat_results.append(first_result)
                await _save_record_to_db(first_result)
                # 检测到链接失效时跳过剩余Cookie，节省时间
                if not first_result.success and "已失效" in first_result.error_msg:
                    task_info.add_log(f"  [{plat_name}] 链接失效，跳过剩余 {len(cookies)-1} 个账号")
                    async with processed_lock:
                        processed_counter["n"] += len(cookies)
                        task_info.processed = processed_counter["n"]
                        task_info.progress = round(processed_counter["n"] / max(total_ops, 1) * 100, 1)
                    continue
            async with processed_lock:
                processed_counter["n"] += 1
                task_info.processed = processed_counter["n"]
                task_info.progress = round(processed_counter["n"] / max(total_ops, 1) * 100, 1)

            # 链接有效，继续用剩余 Cookie 并发举报
            if len(cookies) > 1:
                cookie_tasks = []
                for cid, cstr in cookies[1:]:
                    cookie_tasks.append(_single_report_with_semaphore(
                        semaphore, handler, url, cstr, cid,
                        reason_text, description, task_id, task_info, plat_name,
                    ))
                results = await asyncio.gather(*cookie_tasks, return_exceptions=True)
                for r in results:
                    if isinstance(r, ReportResult):
                        plat_results.append(r)
                        await _save_record_to_db(r)
                    async with processed_lock:
                        processed_counter["n"] += 1
                        task_info.processed = processed_counter["n"]
                        task_info.progress = round(processed_counter["n"] / max(total_ops, 1) * 100, 1)

        return plat_results

    # 所有平台的协程同时启动
    platform_coros = [
        _process_platform(plat, urls_list)
        for plat, urls_list in platform_urls.items()
    ]
    platform_results = await asyncio.gather(*platform_coros, return_exceptions=True)
    for pr in platform_results:
        if isinstance(pr, list):
            all_results.extend(pr)

    # 打包截图为 ZIP
    zip_path = _pack_screenshots(task_id)
    if zip_path:
        task_info.add_log(f"截图已打包: {zip_path}")

    # 生成汇总
    success_count = sum(1 for r in all_results if r.success)
    fail_count = len(all_results) - success_count
    task_info.add_log(
        f"任务完成: 总计 {len(all_results)} 次举报, "
        f"成功 {success_count}, 失败 {fail_count}"
    )

    # 存储结果供 API 层读取
    task_info.result_data = [_result_to_dict(r) for r in all_results]
    task_info.result_file = zip_path or ""
    task_info.message = f"完成: 成功{success_count}/失败{fail_count}"


async def _single_report_with_semaphore(
    semaphore, handler, url, cookie_str, cookie_id,
    reason_text, description, task_id, task_info, plat_name,
) -> ReportResult:
    """并发模式包装：通过 Semaphore 限制最大并发浏览器数"""
    async with semaphore:
        jitter = random.uniform(0, 3)
        await asyncio.sleep(jitter)
        return await _execute_single_report(
            handler, url, cookie_str, cookie_id,
            reason_text, description, task_id, task_info, plat_name,
        )


async def _execute_single_report(
    handler: BaseReport, url: str, cookie_str: str, cookie_id: str,
    reason_text: str, description: str, task_id: str,
    task_info, plat_name: str,
) -> ReportResult:
    """执行单次举报并记录日志，Cookie失效时自动标记"""
    task_info.add_log(f"  [{plat_name}] 账号 {cookie_id} 开始举报...")

    result = await handler.execute(
        url=url, cookie_str=cookie_str, cookie_id=cookie_id,
        reason_text=reason_text, description=description, task_id=task_id,
    )

    if result.success:
        task_info.add_log(
            f"  [{plat_name}] 账号 {cookie_id} 举报成功 ({result.elapsed_sec}s)"
        )
    else:
        task_info.add_log(
            f"  [{plat_name}] 账号 {cookie_id} 举报失败: {result.error_msg} ({result.elapsed_sec}s)"
        )
        # Cookie过期/无效时自动标记失效，避免后续任务继续使用已失效Cookie
        _cookie_fatal_keywords = ["Cookie已过期", "Cookie无效", "需要重新扫码", "登录后举报", "登录即可"]
        if any(kw in result.error_msg for kw in _cookie_fatal_keywords):
            from proxy.cookie_pool import cookie_pool, FailureLevel
            platform = result.platform
            entry = cookie_pool._find_entry(platform, cookie_id)
            if entry:
                cookie_pool._last_used_id[platform] = cookie_id
                cookie_pool.report_failure(platform, FailureLevel.FATAL)
                task_info.add_log(f"  [{plat_name}] 账号 {cookie_id} 已标记为失效")

    return result


def _group_by_platform(urls: List[str]) -> dict[str, List[str]]:
    """按平台分组 URL"""
    groups: dict[str, List[str]] = {}
    for url in urls:
        platform, _ = detect_platform(url)
        groups.setdefault(platform, []).append(url)
    return groups


def _get_cookie_count(platform: str) -> int:
    """获取指定平台有效 Cookie 数量"""
    if platform in getattr(config, "COOKIE_FREE_PLATFORMS", []):
        if not cookie_pool.has_valid_cookie(platform):
            return 1
    return cookie_pool.get_valid_count(platform)


def _get_all_cookies(platform: str) -> List[tuple]:
    """获取指定平台全部有效 Cookie，返回 [(cookie_id, cookie_str), ...]"""
    count = cookie_pool.get_valid_count(platform)
    if count == 0:
        return []
    return cookie_pool.allocate_cookies(platform, count)


def _resolve_reason(platform: str, reason: str) -> str:
    """将用户选择的理由映射为平台举报弹窗中的选择器文本"""
    from media_platform.report.report_config import PLATFORM_REPORT_REASONS
    platform_reasons = PLATFORM_REPORT_REASONS.get(platform, {})
    if reason in platform_reasons:
        return platform_reasons[reason]
    for key, val in platform_reasons.items():
        if reason in key or key in reason:
            return val
    return DEFAULT_REASON.get(platform, "不实信息")


async def _random_sleep():
    """举报操作间的随机延迟（防风控）"""
    interval = getattr(config, "REPORT_INTERVAL_SEC", (5, 10))
    base = random.uniform(interval[0], interval[1])
    jitter_ratio = getattr(config, "SLEEP_JITTER_RATIO", 0.3)
    jitter = base * random.uniform(-jitter_ratio, jitter_ratio)
    sleep_time = max(1, base + jitter)
    await asyncio.sleep(sleep_time)


def _pack_screenshots(task_id: str) -> Optional[str]:
    """将任务截图目录打包为 ZIP 文件"""
    screenshot_dir = os.path.join("data", "report_screenshots", task_id)
    if not os.path.exists(screenshot_dir):
        return None

    files = [f for f in os.listdir(screenshot_dir) if f.endswith(".png")]
    if not files:
        return None

    zip_path = os.path.join("data", "report_screenshots", f"{task_id}.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in files:
            zf.write(os.path.join(screenshot_dir, f), f)

    utils.logger.info(f"[ReportEngine] 截图打包完成: {zip_path} ({len(files)} 张)")
    return zip_path


async def _save_record_to_db(result: ReportResult):
    """将单次举报记录写入 MySQL（静默失败，不阻断主流程）"""
    try:
        from database.external_db import external_db
        await external_db.ensure_pool()
        sql = (
            "INSERT INTO report_records "
            "(task_id, url, platform, cookie_id, reason, success, error_msg, "
            "screenshot_pre_path, screenshot_post_path) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"
        )
        # 从截图路径中解析 task_id
        task_id = ""
        path = result.screenshot_pre_path or result.screenshot_post_path
        if path:
            parts = path.replace("\\", "/").split("/")
            if len(parts) >= 3:
                task_id = parts[-2]

        async with external_db._pool.acquire() as conn:
            async with conn.cursor() as cur:
                await cur.execute(sql, (
                    task_id, result.url, result.platform, result.cookie_id,
                    result.reason, 1 if result.success else 0,
                    result.error_msg, result.screenshot_pre_path,
                    result.screenshot_post_path,
                ))
    except Exception as e:
        utils.logger.warning(f"[ReportEngine] DB写入失败(非致命): {e}")


def _result_to_dict(r: ReportResult) -> dict:
    """将 ReportResult 转为可序列化的字典"""
    return {
        "url": r.url,
        "platform": r.platform,
        "cookie_id": r.cookie_id,
        "reason": r.reason,
        "success": r.success,
        "error_msg": r.error_msg,
        "screenshot_pre_path": r.screenshot_pre_path,
        "screenshot_post_path": r.screenshot_post_path,
        "elapsed_sec": r.elapsed_sec,
    }


def get_latest_screenshot_base64(task_id: str) -> Optional[str]:
    """获取任务最新一张截图的 base64 编码（前端实时预览用）"""
    screenshot_dir = os.path.join("data", "report_screenshots", task_id)
    if not os.path.exists(screenshot_dir):
        return None

    files = sorted(
        [f for f in os.listdir(screenshot_dir) if f.endswith(".png")],
        key=lambda f: os.path.getmtime(os.path.join(screenshot_dir, f)),
        reverse=True,
    )
    if not files:
        return None

    filepath = os.path.join(screenshot_dir, files[0])
    with open(filepath, "rb") as f:
        return base64.b64encode(f.read()).decode()


def get_screenshot_base64(task_id: str, filename: str) -> Optional[str]:
    """获取指定截图文件的 base64 编码"""
    screenshot_dir = os.path.join("data", "report_screenshots", task_id)
    filepath = os.path.join(screenshot_dir, filename)
    if not os.path.exists(filepath):
        return None
    with open(filepath, "rb") as f:
        return base64.b64encode(f.read()).decode()


def generate_report_excel(task_id: str, results: list[dict]) -> Optional[str]:
    """
    生成举报结果 Excel 报告。
    列：序号、链接、平台、账号、举报理由、结果、失败原因、耗时、提交前截图、提交后截图
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        utils.logger.error("[ReportEngine] openpyxl 未安装，无法生成 Excel")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "举报结果"

    headers = ["序号", "链接", "平台", "账号", "举报理由", "结果", "失败原因", "耗时(s)",
               "提交前截图", "提交后截图"]
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 结果行的成功/失败高亮
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for idx, r in enumerate(results, 1):
        row = idx + 1
        is_success = r.get("success", False)
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=r.get("url", ""))
        ws.cell(row=row, column=3, value=PLATFORM_NAMES.get(r.get("platform", ""), r.get("platform", "")))
        ws.cell(row=row, column=4, value=r.get("cookie_id", ""))
        ws.cell(row=row, column=5, value=r.get("reason", ""))

        result_cell = ws.cell(row=row, column=6, value="成功" if is_success else "失败")
        result_cell.fill = success_fill if is_success else fail_fill
        result_cell.font = Font(bold=True)

        ws.cell(row=row, column=7, value=r.get("error_msg", ""))
        ws.cell(row=row, column=8, value=r.get("elapsed_sec", 0))
        ws.cell(row=row, column=9, value=os.path.basename(r.get("screenshot_pre_path", "")))
        ws.cell(row=row, column=10, value=os.path.basename(r.get("screenshot_post_path", "")))

    # 列宽调整
    col_widths = [6, 60, 10, 15, 12, 8, 40, 10, 30, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    excel_dir = os.path.join("data", "report_screenshots")
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = os.path.join(excel_dir, f"{task_id}_report.xlsx")
    wb.save(excel_path)
    utils.logger.info(f"[ReportEngine] Excel报告已生成: {excel_path}")
    return excel_path
