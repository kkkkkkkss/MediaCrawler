# -*- coding: utf-8 -*-
# 评论数据导出模块
# 将 TaskInfo.comments_data 导出为 JSON 或 Excel 文件

import json
import pathlib
from datetime import datetime
from typing import Dict, List, Optional

from tools import utils

_PLATFORM_NAMES = {
    "dy": "抖音", "ks": "快手", "bili": "B站",
    "toutiao": "今日头条", "xhs": "小红书", "wb": "微博",
}


def export_comments(
    comments_data: List[Dict],
    task_id: str = "",
    format: str = "json",
) -> str:
    """
    导出评论数据为文件。

    Args:
        comments_data: TaskInfo.comments_data，按作品分组的评论列表
        task_id: 任务 ID
        format: "json" 或 "excel"

    Returns:
        生成的文件路径
    """
    if format == "excel":
        return _export_excel(comments_data, task_id)
    return _export_json(comments_data, task_id)


def _export_json(comments_data: List[Dict], task_id: str) -> str:
    """导出为 JSON 文件"""
    output_dir = pathlib.Path("data/url_check/comments")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{task_id}_comments.json"

    total = sum(len(item.get("comments", [])) for item in comments_data)
    data = {
        "task_id": task_id,
        "total_comments": total,
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "results": comments_data,
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    utils.logger.info(f"[comment_export] JSON 评论导出: {output_path}")
    return str(output_path)


def _export_excel(comments_data: List[Dict], task_id: str) -> str:
    """导出为 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl 未安装，请执行: pip install openpyxl")

    output_dir = pathlib.Path("data/url_check/comments")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{task_id}_comments.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "评论数据"

    # 表头
    headers = [
        "作品URL", "平台", "评论ID", "评论作者",
        "评论内容", "评论点赞数", "回复数", "评论时间",
    ]
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 写入数据
    row_idx = 2
    for item in comments_data:
        content_url = item.get("content_url", "")
        platform = item.get("platform", "")
        plat_name = _PLATFORM_NAMES.get(platform, platform)

        for comment in item.get("comments", []):
            values = [
                content_url,
                plat_name,
                comment.get("comment_id", ""),
                comment.get("author_name", ""),
                comment.get("comment_text", ""),
                comment.get("comment_like_count", ""),
                comment.get("comment_reply_count", ""),
                str(comment.get("comment_time", "")) if comment.get("comment_time") else "",
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
            row_idx += 1

    # 自动列宽
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 60)

    wb.save(str(output_path))
    utils.logger.info(f"[comment_export] Excel 评论导出: {output_path}")
    return str(output_path)


def comments_to_json_data(comments_data: List[Dict], task_id: str = "") -> Dict:
    """将评论数据转为标准化 JSON 结构（用于回调）"""
    total = sum(len(item.get("comments", [])) for item in comments_data)
    return {
        "task_id": task_id,
        "total_comments": total,
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "results": comments_data,
    }
