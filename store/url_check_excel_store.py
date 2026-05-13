# -*- coding: utf-8 -*-
# url_check 模式专用 Excel 报表输出
# 将处理结果输出为包含中文列名的格式化 Excel 文件

import pathlib
from datetime import datetime
from typing import Any, Dict, List, Optional

from tools import utils

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# 平台代码 → 中文名映射
_PLATFORM_NAMES = {
    "dy": "抖音",
    "ks": "快手",
    "bili": "B站",
    "toutiao": "今日头条",
    "xhs": "小红书",
    "wb": "微博",
    "tieba": "百度贴吧",
    "zhihu": "知乎",
    "unknown": "未知",
}

# 平台代码 → 默认内容类型映射
_PLATFORM_DEFAULT_TYPE = {
    "dy": "视频",
    "ks": "视频",
    "bili": "视频",
    "toutiao": "视频",
    "xhs": "笔记",
    "wb": "微博",
    "tieba": "帖子",
    "zhihu": "回答",
}

# Excel 列定义：(英文key, 中文列名)
_COLUMNS = [
    ("id", "序号"),
    ("type", "内容类型"),
    ("web_name", "平台名称"),
    ("title", "标题"),
    ("author", "作者"),
    ("praise_count", "点赞数"),
    ("reply_count", "评论数"),
    ("visit_count", "播放/浏览数"),
    ("share_count", "分享数"),
    ("is_valid", "链接有效性"),
    ("url", "原始链接"),
]

_MAX_TITLE_LENGTH = 50


def _truncate_title(value: Any) -> str:
    """标题输出统一限制在 _MAX_TITLE_LENGTH 字数以内。"""
    if value is None:
        return ""
    title = str(value).strip()
    if len(title) <= _MAX_TITLE_LENGTH:
        return title
    return title[:_MAX_TITLE_LENGTH]


def _detect_content_type(platform: str, raw_json: Optional[Dict]) -> str:
    """根据平台和原始数据判断内容类型"""
    if not raw_json or not isinstance(raw_json, dict):
        return _PLATFORM_DEFAULT_TYPE.get(platform, "未知")

    if platform == "toutiao":
        if raw_json.get("video_play_count") or raw_json.get("video_duration"):
            return "视频"
        article_type = raw_json.get("article_type") or raw_json.get("type")
        if article_type and "video" in str(article_type).lower():
            return "视频"
        return "文章"

    if platform == "xhs":
        note_type = raw_json.get("type")
        if note_type == "video":
            return "视频"
        return "图文笔记"

    return _PLATFORM_DEFAULT_TYPE.get(platform, "未知")


def _extract_author(platform: str, raw_json: Optional[Dict], metrics: Optional[Dict]) -> str:
    """从原始数据中提取作者名"""
    if metrics and metrics.get("author"):
        return str(metrics["author"])

    if not raw_json or not isinstance(raw_json, dict):
        return ""

    # 各平台作者字段路径
    paths = {
        "dy": [("author", "nickname"), ("author_info", "nickname")],
        "bili": [("View", "owner", "name"), ("owner", "name")],
        "ks": [("photo", "userName"), ("userName",), ("author", "name")],
        "toutiao": [("source",), ("media_name",), ("author", "name")],
        "xhs": [("user", "nickname"), ("note_user", "nickname")],
        "wb": [("mblog", "user", "screen_name"), ("user", "screen_name")],
    }

    for path in paths.get(platform, []):
        obj = raw_json
        for key in path:
            if isinstance(obj, dict):
                obj = obj.get(key)
            else:
                obj = None
                break
        if obj and isinstance(obj, str):
            return obj

    return ""


def _extract_title(platform: str, raw_json: Optional[Dict], metrics: Optional[Dict], row_title: str = "") -> str:
    """从原始数据或已提取的 metrics 中获取标题/正文摘要"""
    # 优先使用 DOM 提取的标题（如头条从页面h1标签获取）
    if row_title:
        return _truncate_title(row_title)
    if metrics and metrics.get("title"):
        return _truncate_title(metrics["title"])

    if not raw_json or not isinstance(raw_json, dict):
        return ""

    # 各平台标题提取逻辑
    if platform == "wb":
        from tools.fallback_field_map import _extract_wb_title
        return _truncate_title(_extract_wb_title(raw_json))

    if platform == "dy":
        # 抖音: aweme_detail.desc / desc / title
        detail = raw_json.get("aweme_detail") or raw_json
        for key in ("desc", "title", "share_info.share_title"):
            val = detail.get(key)
            if val and isinstance(val, str) and len(val) > 1:
                return _truncate_title(val)
        share_info = detail.get("share_info", {})
        if isinstance(share_info, dict):
            st = share_info.get("share_title", "")
            if st:
                return _truncate_title(st)

    if platform == "bili":
        # B站: title / data.title / data.View.title
        for key in ("title",):
            if raw_json.get(key):
                return _truncate_title(raw_json[key])
        data = raw_json.get("data", {})
        if isinstance(data, dict):
            if data.get("title"):
                return _truncate_title(data["title"])
            view = data.get("View", {})
            if isinstance(view, dict) and view.get("title"):
                return _truncate_title(view["title"])

    if platform == "ks":
        # 快手: visionVideoDetail.photo.caption / caption
        detail = raw_json.get("visionVideoDetail", raw_json)
        if isinstance(detail, dict):
            photo = detail.get("photo", detail)
            if isinstance(photo, dict):
                cap = photo.get("caption", "")
                if cap:
                    return _truncate_title(cap)

    if platform == "toutiao":
        # 头条: title / articleInfo.title / data.title
        for key in ("title",):
            if raw_json.get(key):
                return _truncate_title(raw_json[key])
        for sub_key in ("articleInfo", "data", "itemInfo"):
            sub = raw_json.get(sub_key, {})
            if isinstance(sub, dict) and sub.get("title"):
                return _truncate_title(sub["title"])

    if platform == "xhs":
        # 小红书: title / note.title
        if raw_json.get("title"):
            return _truncate_title(raw_json["title"])
        note = raw_json.get("note", {})
        if isinstance(note, dict) and note.get("title"):
            return _truncate_title(note["title"])

    return ""


def generate_url_check_excel(results: List[Dict], output_path: Optional[str] = None) -> str:
    """
    将 url_check 处理结果输出为 Excel 文件。

    Args:
        results: UrlCheckCrawler._all_results 列表
        output_path: 输出路径，为空则自动生成

    Returns:
        生成的 Excel 文件路径
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl 未安装，请执行: pip install openpyxl")

    # 生成输出路径
    if not output_path:
        base_dir = pathlib.Path("data/url_check/excel")
        base_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(base_dir / f"url_check_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "URL检测结果"

    # 写入表头
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_idx, (_, cn_name) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=cn_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # 写入数据行
    for row_idx, result in enumerate(results, 2):
        platform = result.get("_platform", "unknown")
        metrics = result.get("_metrics", {})
        raw_json = result.get("_raw_json")
        is_valid = result.get("_is_valid", 0)

        row_data = {
            "id": result.get("id", row_idx - 1),
            "type": _detect_content_type(platform, raw_json),
            "web_name": _PLATFORM_NAMES.get(platform, "未知"),
            "title": _extract_title(platform, raw_json, metrics, result.get("_title", "")),
            "author": _extract_author(platform, raw_json, metrics),
            "praise_count": metrics.get("praise_count", ""),
            "reply_count": metrics.get("reply_count", ""),
            "visit_count": metrics.get("visit_count", ""),
            "share_count": metrics.get("share_count", ""),
            "is_valid": "有效" if is_valid == 1 else "无效",
            "url": result.get("url", ""),
        }

        for col_idx, (key, _) in enumerate(_COLUMNS, 1):
            value = row_data.get(key, "")
            if value is None:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # 自动调整列宽
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 60)

    wb.save(output_path)
    utils.logger.info(f"[url_check_excel] Excel 报表已保存: {output_path}")
    return output_path


# ──────────────────── 合并策略配置 ────────────────────

# 「强制覆盖」列：不论原值是什么，始终用爬取数据覆盖
_OVERWRITE_COLUMNS = {
    "转发数": "share_count",
    "评论数": "reply_count",
    "点赞数": "praise_count",
    "分享数": "share_count",   # "分享数"与"转发数"等效
}

# 「标题」等效列名正则匹配模式（任一匹配即复用）
import re as _re
_TITLE_COLUMN_PATTERNS = [
    _re.compile(r"标题"),       # 含"标题"的列名
    _re.compile(r"^title$", _re.IGNORECASE),
    _re.compile(r"微博内容"),   # "标题/微博内容"等
    _re.compile(r"内容$"),      # "正文内容"等
]

# 标题列中视为"空"的占位符值
_TITLE_EMPTY_PLACEHOLDERS = {"", "-", "\t-", "—", "无", "无标题", "null", "None", "/"}

# 「作者」等效列名：有其中任何一个就复用，不新增
_AUTHOR_COLUMN_NAMES = ["原文作者", "作者", "author", "转发/评论作者"]

# 「平台名称」等效列名
_PLATFORM_COLUMN_NAMES = ["平台名称", "来源网站", "平台", "网站名称"]

# 需要确保存在的列及其数据键（不存在时追加到末尾）
# 顺序为追加顺序，链接有效性始终放最后
_ENSURE_COLUMNS = [
    ("平台名称", "platform"),
    ("标题", "title"),
    ("作者", "author"),
    ("点赞数", "praise_count"),
    ("评论数", "reply_count"),
    ("分享数", "share_count"),
    ("播放/浏览数", "visit_count"),
    ("链接有效性", "is_valid"),
]

_PLATFORM_NAMES = {
    "dy": "抖音", "ks": "快手", "bili": "B站",
    "toutiao": "头条", "xhs": "小红书", "wb": "微博",
}


def _is_empty_value(val) -> bool:
    """判断单元格值是否视为"空"（None/空字符串/纯0）"""
    if val is None:
        return True
    s = str(val).strip()
    return s == "" or s == "0" or s == "0.0"


def _find_equivalent_col(header_map: Dict[str, int], candidates: list) -> Optional[int]:
    """在表头中查找候选列名中第一个匹配的索引"""
    for name in candidates:
        if name in header_map:
            return header_map[name]
    return None


def _find_title_col(header_map: Dict[str, int]) -> Optional[int]:
    """用正则在表头中查找标题等效列"""
    for col_name, idx in header_map.items():
        for pattern in _TITLE_COLUMN_PATTERNS:
            if pattern.search(col_name):
                return idx
    return None


def merge_results_to_excel(
    source_path: str,
    results: List[Dict],
    output_path: str,
    url_column: str = "原文/评论链接",
) -> str:
    """
    在原始 Excel 基础上回填爬取结果，保留全部原始列和数据。

    策略：
    - 赞/评/转(分享)三列：始终用爬取数据覆盖（不论原值）
    - 作者/平台等列：有原始值保留，为空则填充
    - 表中缺失的标准列（点赞数/评论数/分享数/播放浏览数/平台名称/作者/链接有效性）：
      追加到末尾
    - 链接有效性始终是最后一列
    """
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl 未安装，请执行: pip install openpyxl")

    wb = openpyxl.load_workbook(source_path)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}

    # ─── 定位 URL 列 ───
    url_col_idx = None
    for name, idx in header_map.items():
        if name.lower() == url_column.lower():
            url_col_idx = idx
            break
    if url_col_idx is None:
        for name, idx in header_map.items():
            if "链接" in name or "url" in name.lower() or "地址" in name:
                url_col_idx = idx
                break
    if url_col_idx is None:
        utils.logger.warning(f"[merge_excel] 未找到URL列'{url_column}'，跳过合并")
        return generate_url_check_excel(results, output_path)

    # ─── 建立 url → result 索引 ───
    result_map: Dict[str, Dict] = {}
    for r in results:
        url = (r.get("url") or "").strip()
        if url:
            result_map[url] = r

    # ─── 定位「强制覆盖」列 ───
    overwrite_col_indices: Dict[str, int] = {}
    for cn_name, _ in _OVERWRITE_COLUMNS.items():
        if cn_name in header_map:
            overwrite_col_indices[cn_name] = header_map[cn_name]

    # ─── 定位已有的标题列、作者列和平台列 ───
    title_col_idx = _find_title_col(header_map)
    author_col_idx = _find_equivalent_col(header_map, _AUTHOR_COLUMN_NAMES)
    platform_col_idx = _find_equivalent_col(header_map, _PLATFORM_COLUMN_NAMES)

    # ─── 计算需要追加的缺失列 ───
    append_cols: List[tuple] = []
    next_col = len(headers)

    for cn_name, key in _ENSURE_COLUMNS:
        already_exists = False

        if key == "title":
            already_exists = title_col_idx is not None
        elif key == "author":
            already_exists = author_col_idx is not None
        elif key == "platform":
            already_exists = platform_col_idx is not None
        elif key == "praise_count":
            already_exists = "点赞数" in header_map
        elif key == "reply_count":
            already_exists = "评论数" in header_map
        elif key == "share_count":
            already_exists = ("转发数" in header_map or "分享数" in header_map)
        elif key == "visit_count":
            already_exists = "播放/浏览数" in header_map
        elif key == "is_valid":
            already_exists = "链接有效性" in header_map

        if not already_exists:
            append_cols.append((cn_name, key, next_col))
            ws.cell(row=1, column=next_col + 1, value=cn_name)
            next_col += 1

    # ─── 逐行回填数据 ───
    for row_idx in range(2, ws.max_row + 1):
        url_cell = ws.cell(row=row_idx, column=url_col_idx + 1).value
        if not url_cell:
            continue
        url_str = str(url_cell).strip()
        result = result_map.get(url_str)
        if not result:
            continue

        metrics = result.get("_metrics", {})
        is_valid = result.get("_is_valid", 0)
        raw_json = result.get("_raw_json")
        platform = result.get("_platform", "unknown")

        # ① 强制覆盖列（赞/评/转）：始终用爬取值覆盖
        for cn_name, col_idx in overwrite_col_indices.items():
            metric_key = _OVERWRITE_COLUMNS[cn_name]
            new_val = metrics.get(metric_key)
            if new_val is not None:
                ws.cell(row=row_idx, column=col_idx + 1, value=new_val)

        # ② 标题列：空值或占位符时填充（覆盖 "-"、"无" 等无效标题）
        if title_col_idx is not None:
            existing = ws.cell(row=row_idx, column=title_col_idx + 1).value
            existing_str = str(existing).strip() if existing is not None else ""
            if existing_str in _TITLE_EMPTY_PLACEHOLDERS:
                title_val = _extract_title(platform, raw_json, metrics, result.get("_title", ""))
                if title_val:
                    ws.cell(row=row_idx, column=title_col_idx + 1, value=title_val)
            elif len(existing_str) > _MAX_TITLE_LENGTH:
                ws.cell(
                    row=row_idx,
                    column=title_col_idx + 1,
                    value=_truncate_title(existing_str),
                )

        # ③ 作者列：有值保留，空值填充
        if author_col_idx is not None:
            existing = ws.cell(row=row_idx, column=author_col_idx + 1).value
            if _is_empty_value(existing):
                author_val = _extract_author(platform, raw_json, metrics)
                if author_val:
                    ws.cell(row=row_idx, column=author_col_idx + 1, value=author_val)

        # ④ 平台列：有值保留，空值填充
        if platform_col_idx is not None:
            existing = ws.cell(row=row_idx, column=platform_col_idx + 1).value
            if _is_empty_value(existing):
                ws.cell(
                    row=row_idx, column=platform_col_idx + 1,
                    value=_PLATFORM_NAMES.get(platform, platform)
                )

        # ⑤ 追加的新列
        for cn_name, key, col_idx in append_cols:
            if key == "is_valid":
                val = "有效" if is_valid == 1 else "无效"
            elif key == "visit_count":
                val = metrics.get("visit_count", "")
            elif key == "title":
                val = _extract_title(platform, raw_json, metrics, result.get("_title", ""))
            elif key == "author":
                val = _extract_author(platform, raw_json, metrics)
            elif key == "platform":
                val = _PLATFORM_NAMES.get(platform, platform)
            elif key in ("praise_count", "reply_count", "share_count"):
                val = metrics.get(key, "")
            else:
                val = metrics.get(key, "")
            if val is None:
                val = ""
            ws.cell(row=row_idx, column=col_idx + 1, value=val)

    # 保存
    pathlib.Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    utils.logger.info(f"[merge_excel] 已合并结果到原始 Excel: {output_path}")
    return output_path
